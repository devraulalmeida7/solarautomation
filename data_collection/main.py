import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

class ExcelSumApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Somatório Excel")
        self.file_path = None
        
        # Configuração da interface
        self.main_frame = tk.Frame(root, padx=20, pady=20)
        self.main_frame.pack()
        
        # Botão para carregar arquivo
        self.load_btn = tk.Button(
            self.main_frame,
            text="Carregar Arquivo Excel",
            command=self.load_file,
            width=30
        )
        self.load_btn.pack(pady=10)
        
        # Botão para somar todas as colunas (sem filtro)
        self.sum_all_btn = tk.Button(
            self.main_frame,
            text="Somar Todas as Colunas",
            command=self.sum_all_columns,
            state=tk.DISABLED,
            width=30
        )
        self.sum_all_btn.pack(pady=5)
        
        # Botão para somar colunas com filtro de horário
        self.sum_time_btn = tk.Button(
            self.main_frame,
            text="Somar Colunas (10h-15h)",
            command=self.sum_columns_time_filter,
            state=tk.DISABLED,
            width=30
        )
        self.sum_time_btn.pack(pady=5)
        
        # Status
        self.status_var = tk.StringVar()
        self.status_label = tk.Label(
            self.main_frame,
            textvariable=self.status_var,
            fg="blue"
        )
        self.status_label.pack(pady=10)
    
    def load_file(self):
        """Carrega o arquivo Excel selecionado pelo usuário."""
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if self.file_path:
            self.status_var.set(f"Arquivo carregado: {self.file_path}")
            self.sum_all_btn.config(state=tk.NORMAL)
            self.sum_time_btn.config(state=tk.NORMAL)
        else:
            self.status_var.set("Nenhum arquivo selecionado.")
    
    def sum_all_columns(self):
        """Soma todos os valores numéricos de cada coluna (a partir da coluna B)."""
        if not self.file_path:
            messagebox.showerror("Erro", "Nenhum arquivo carregado!")
            return
        
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            resultados = {}
            
            # Percorre cada coluna a partir da B
            for col in range(2, ws.max_column + 1):
                col_letter = chr(64 + col)  # B, C, D...
                soma = 0
                
                # Soma todos os valores numéricos da coluna
                for linha in range(2, ws.max_row + 1):
                    valor = ws[f"{col_letter}{linha}"].value
                    if isinstance(valor, (int, float)):
                        soma += valor
                
                resultados[col_letter] = soma
            
            # Adiciona os resultados no final da planilha
            ultima_linha = ws.max_row + 1
            ws[f"A{ultima_linha}"] = "Somas totais por coluna:"
            for col_letter, soma in resultados.items():
                ws[f"{col_letter}{ultima_linha}"] = soma
            
            wb.save(self.file_path)
            
            # Exibe resumo
            mensagem = "Somas totais por coluna:\n"
            for col_letter, soma in resultados.items():
                mensagem += f"\nColuna {col_letter}: {soma}"
            
            messagebox.showinfo("Resultados", mensagem)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao processar:\n{str(e)}")
    
    def sum_columns_time_filter(self):
        """Soma os valores de cada coluna, filtrando por horário na coluna A (10h-15h)."""
        if not self.file_path:
            messagebox.showerror("Erro", "Nenhum arquivo carregado!")
            return
        
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            resultados = {}
            
            # Percorre cada coluna a partir da B
            for col in range(2, ws.max_column + 1):
                col_letter = chr(64 + col)  # B, C, D...
                soma = 0
                linhas_validas = []
                
                # Verifica o horário na coluna A e soma os valores correspondentes
                for linha in range(2, ws.max_row + 1):
                    hora = ws[f"A{linha}"].value
                    valor = ws[f"{col_letter}{linha}"].value
                    
                    if hora and isinstance(valor, (int, float)):
                        hora_str = str(hora)
                        if "10:00:00Z" <= hora_str[-9:] <= "15:00:00Z":
                            soma += valor
                            linhas_validas.append((linha, hora_str, valor))
                
                resultados[col_letter] = {
                    "soma": soma,
                    "detalhes": linhas_validas
                }
            
            # Adiciona os resultados no final da planilha
            ultima_linha = ws.max_row + 1
            ws[f"A{ultima_linha}"] = "Somas (10h-15h) por coluna:"
            for col_letter, dados in resultados.items():
                ws[f"{col_letter}{ultima_linha}"] = dados["soma"]
            
            wb.save(self.file_path)
            
            # Exibe resumo
            mensagem = "Somas por coluna (10h-15h):\n"
            for col_letter, dados in resultados.items():
                mensagem += f"\nColuna {col_letter}: {dados['soma']}"
            
            messagebox.showinfo("Resultados", mensagem)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao processar:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSumApp(root)
    root.mainloop()