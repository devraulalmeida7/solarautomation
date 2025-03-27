# Adicionado dados a uma planilha existente.
from functions.append import average
from openpyxl import load_workbook

wb = load_workbook('D:/coding/2025/solarproject/data_collection/tests/dados16.xlsx')

sheet = wb['page1']

# average()

last_line = sheet.max_row - 4
last_column = sheet.max_column

last_cell_value = sheet.cell(row=last_line,column=4).value

print(last_cell_value)

wb.save('D:/coding/2025/solarproject/data_collection/tests/dados16.xlsx')